"""xlsx 渲染器：ArtifactSpec → 公式驱动的回测底稿。

"可交互"的实现方式：多资产日期对齐由 Python 完成（确定性、可测），
收益/回撤/滚动波动率/相关性全部写成 Excel 公式并引用"参数"sheet——
评审者改参数单元格（如滚动窗口）即可联动重算，不是写死的数值。
"""

from __future__ import annotations

import io
from collections.abc import Mapping
from datetime import UTC, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from finance_agent.artifacts.spec import (
    ArtifactSpec,
    DataSheetBlock,
    HeadingBlock,
    MetricsSheetBlock,
    NarrativeBlock,
    TableBlock,
)
from finance_agent.provenance import Evidence
from finance_agent.skills.loader import SkillInfo

SUPPORTED_BLOCKS = {"heading", "narrative", "table", "data_sheet", "metrics_sheet"}

_BOLD = Font(bold=True)


class RenderError(RuntimeError):
    pass


# 占位符黑名单：模型在 table 里写"公式"等占位、指望渲染器代填的失败模式
# （真实事故：年度收益/核心指标两个 sheet 全是字面量"公式"）
_PLACEHOLDER_CELLS = {"公式", "待填", "待补", "占位", "tbd", "todo", "placeholder"}


def _reject_placeholder_tables(spec: ArtifactSpec) -> None:
    for block in spec.blocks:
        if not isinstance(block, TableBlock):
            continue
        bad = [
            str(cell).strip()
            for row in block.rows for cell in row
            if str(cell).strip().lower() in _PLACEHOLDER_CELLS
        ]
        if bad:
            raise RenderError(
                f"table「{block.caption or '未命名'}」含 {len(bad)} 个占位符单元格"
                f"（如 {bad[0]!r}）。可计算指标不要用 table 手写：年度收益与核心风险"
                "收益指标由 metrics_sheet 自动生成（「年度收益」「汇总」sheet，全公式）；"
                "table 仅承载真实的定性内容（事件清单、口径对照等）。"
                "请删除该 table 或改为真实内容后重试。"
            )


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _BOLD


def _data_sheet(wb: Workbook, block: DataSheetBlock, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(block.sheet_name[:31])
    _write_header(ws, ["日期", "开盘", "最高", "最低", "收盘", "成交量"])
    for row in df.itertuples(index=False):
        ws.append([row.date, row.open, row.high, row.low, row.close, row.volume])
    ws.freeze_panes = "A2"


def _align(dfs: list[pd.DataFrame]) -> pd.DataFrame:
    """多资产按日期内连接（收盘价）。加密资产 7 天交易、股票 5 天——取交集。"""
    aligned = dfs[0][["date", "close"]].rename(columns={"close": "close_0"})
    for i, df in enumerate(dfs[1:], start=1):
        aligned = aligned.merge(
            df[["date", "close"]].rename(columns={"close": f"close_{i}"}),
            on="date", how="inner",
        )
    return aligned.sort_values("date").reset_index(drop=True)


def _metrics_sheets(
    wb: Workbook, block: MetricsSheetBlock, datasets: Mapping[str, pd.DataFrame]
) -> None:
    labels = block.labels or block.data_refs
    if len(labels) != len(block.data_refs):
        raise RenderError("metrics_sheet 的 labels 数量须与 data_refs 一致")
    dfs = [datasets[ref] for ref in block.data_refs]
    aligned = _align(dfs)
    if len(aligned) < block.rolling_window + 2:
        raise RenderError(
            f"对齐后数据不足：{len(aligned)} 行 < 滚动窗口 {block.rolling_window}+2"
        )
    n_assets = len(dfs)
    n_rows = len(aligned)
    last = n_rows + 1  # 数据末行（含表头偏移）

    # 参数 sheet：公式的唯一参数来源，改这里即联动
    params = wb.create_sheet("参数")
    params.append(["参数", "值", "说明"])
    for cell in params[1]:
        cell.font = _BOLD
    params.append(["滚动窗口（交易日）", block.rolling_window, "改动后指标sheet的滚动波动率联动重算"])
    params.append(["年化系数", block.annualization, "波动率年化用，√年化系数"])

    # 对齐 sheet：确定性代码产出的输入数据
    al = wb.create_sheet("对齐")
    _write_header(al, ["日期"] + [f"{label}收盘" for label in labels])
    for row in aligned.itertuples(index=False):
        al.append(list(row))
    al.freeze_panes = "A2"

    # 指标 sheet：全公式
    mx = wb.create_sheet("指标")
    headers = ["日期"]
    for label in labels:
        headers += [f"{label}日收益", f"{label}净值(=100)", f"{label}回撤", f"{label}滚动年化波动"]
    _write_header(mx, headers)
    for i in range(2, last + 1):
        row: list = [f"='对齐'!A{i}"]
        for a in range(n_assets):
            src = get_column_letter(2 + a)          # 对齐 sheet 的收盘列
            ret_col = get_column_letter(2 + a * 4)   # 本 sheet 日收益列
            row += [
                "" if i == 2 else f"='对齐'!{src}{i}/'对齐'!{src}{i - 1}-1",
                f"='对齐'!{src}{i}/'对齐'!{src}$2*100",
                f"='对齐'!{src}{i}/MAX('对齐'!{src}$2:'对齐'!{src}{i})-1",
                (
                    f"=IF(ROW()-2>='参数'!$B$2,"
                    f"STDEV(OFFSET({ret_col}{i},1-'参数'!$B$2,0,'参数'!$B$2,1))"
                    f"*SQRT('参数'!$B$3),\"\")"
                ),
            ]
        mx.append(row)
    mx.freeze_panes = "A2"

    # 汇总 sheet：区间核心指标（同样全公式）。行号被夏普公式引用，用变量跟踪
    sm = wb.create_sheet("汇总")
    _write_header(sm, ["指标"] + list(labels))
    def _cols(offset: int) -> list[str]:
        return [get_column_letter(2 + a * 4 + offset) for a in range(n_assets)]
    def _al_cols() -> list[str]:
        return [get_column_letter(2 + a) for a in range(n_assets)]

    _PCT = "0.00%"

    def _sm_row(name: str, formulas: list[str], fmt: str | None = _PCT) -> int:
        sm.append([name] + formulas)
        r = sm.max_row
        if fmt:
            for cell in sm[r][1:]:
                cell.number_format = fmt
        return r

    n_ret = last - 2   # 日收益样本数（首行无收益）
    _sm_row("区间总收益", [f"='对齐'!{c}{last}/'对齐'!{c}2-1" for c in _al_cols()])
    r_cagr = _sm_row("年化收益率（CAGR）", [
        f"=POWER('对齐'!{c}{last}/'对齐'!{c}2,'参数'!$B$3/{n_ret})-1" for c in _al_cols()
    ])
    r_vol = _sm_row("年化波动率（全样本）", [
        f"=STDEV('指标'!{c}3:{c}{last})*SQRT('参数'!$B$3)" for c in _cols(0)
    ])
    _sm_row("夏普比率（rf=0）", [
        f"={get_column_letter(2 + a)}{r_cagr}/{get_column_letter(2 + a)}{r_vol}"
        for a in range(n_assets)
    ], fmt="0.00")
    _sm_row("最大回撤", [f"=MIN('指标'!{c}3:{c}{last})" for c in _cols(2)])
    _sm_row("正收益天数占比", [
        f'=COUNTIF(\'指标\'!{c}3:{c}{last},">0")/COUNT(\'指标\'!{c}3:{c}{last})'
        for c in _cols(0)
    ])
    if n_assets == 2:
        c0, c1 = _cols(0)
        _sm_row("日收益相关系数",
                [f"=CORREL('指标'!{c0}3:{c0}{last},'指标'!{c1}3:{c1}{last})", ""],
                fmt="0.00")

    # 年度收益 sheet：按对齐数据的真实年份边界生成公式（确定性代码定行号）。
    # 真实事故：模型想给"年度收益对比"却只能在 table 里写"公式"占位——
    # 可计算内容必须是渲染器能力，不靠 LLM 手写。
    yr = wb.create_sheet("年度收益")
    headers = ["年份"] + [f"{label}收益" for label in labels]
    if n_assets == 2:
        headers.append(f"差额（{labels[1]}−{labels[0]}）")
    _write_header(yr, headers)
    dates = aligned["date"].astype(str)
    year_bounds: list[tuple[str, int, int]] = []   # (年, 首行, 末行)——sheet 行号
    for year in sorted(dates.str[:4].unique()):
        idx = aligned.index[dates.str[:4] == year]
        year_bounds.append((year, int(idx[0]) + 2, int(idx[-1]) + 2))
    prev_end: int | None = None
    for year, _first, year_end in year_bounds:
        base = 2 if prev_end is None else prev_end   # 上一年最后共同交易日收盘为基准
        row: list = [year] + [f"='对齐'!{c}{year_end}/'对齐'!{c}{base}-1" for c in _al_cols()]
        if n_assets == 2:
            r = yr.max_row + 1
            row.append(f"=C{r}-B{r}")
        yr.append(row)
        prev_end = year_end
    full: list = ["全周期"] + [f"='对齐'!{c}{last}/'对齐'!{c}2-1" for c in _al_cols()]
    if n_assets == 2:
        r = yr.max_row + 1
        full.append(f"=C{r}-B{r}")
    yr.append(full)
    for cells in yr.iter_rows(min_row=2, max_row=yr.max_row, min_col=2):
        for cell in cells:
            cell.number_format = _PCT
    yr.append([])
    yr.append(["注：年度收益以上一年最后共同交易日收盘为基准（首年度以区间首日为基准）；"
               "首末年度可能为区间内的部分年度。"])

    # 图表 sheet：归一化净值曲线
    chart_ws = wb.create_sheet("图表")
    chart = LineChart()
    chart.title = "归一化净值（起点=100）"
    chart.height, chart.width = 12, 24
    dates_ref = Reference(mx, min_col=1, min_row=2, max_row=last)
    for a, _label in enumerate(labels):
        col = 2 + a * 4 + 1  # 净值列
        series_ref = Reference(mx, min_col=col, min_row=1, max_row=last)
        chart.add_data(series_ref, titles_from_data=True)
    chart.set_categories(dates_ref)
    chart_ws.add_chart(chart, "B2")


def _notes_sheet(wb: Workbook, blocks: list) -> None:
    notes = [b for b in blocks if isinstance(b, (HeadingBlock, NarrativeBlock))]
    if not notes:
        return
    ws = wb.create_sheet("说明", 0)
    ws.column_dimensions["A"].width = 110
    for block in notes:
        if isinstance(block, HeadingBlock):
            ws.append([block.text])
            ws[f"A{ws.max_row}"].font = _BOLD
        else:
            for para in block.text.split("\n\n"):
                ws.append([para.strip()])
            if block.evidence_refs:
                ws.append([f"［溯源：{'、'.join(block.evidence_refs)}］"])
        ws.append([])


def _table_sheet(wb: Workbook, block: TableBlock, index: int) -> None:
    name = (block.caption or f"表{index}")[:31]
    ws = wb.create_sheet(name)
    _write_header(ws, block.headers)
    for row in block.rows:
        ws.append(row)
    if block.evidence_refs:
        ws.append([])
        ws.append([f"溯源：{'、'.join(block.evidence_refs)}"])


def _evidence_sheet(wb: Workbook, evidence: Mapping[str, Evidence]) -> None:
    ws = wb.create_sheet("溯源")
    _write_header(ws, ["evidence_id", "类型", "来源URL", "抓取时间", "查询", "摘录"])
    for ev in evidence.values():
        ws.append([ev.id, ev.kind, ev.source_url, ev.fetched_at, str(ev.query), ev.excerpt])
    ws.column_dimensions["C"].width = 60


def render_xlsx(
    spec: ArtifactSpec,
    *,
    datasets: Mapping[str, pd.DataFrame],
    evidence: Mapping[str, Evidence],
    skill: SkillInfo,
) -> bytes:
    if spec.kind != "xlsx":
        raise RenderError(f"kind 不匹配：期望 xlsx，得到 {spec.kind}")
    unsupported = {b.type for b in spec.blocks} - SUPPORTED_BLOCKS
    if unsupported:
        raise RenderError(f"xlsx 渲染器不支持 block 类型：{sorted(unsupported)}")
    _reject_placeholder_tables(spec)
    data_blocks = [b for b in spec.blocks if isinstance(b, DataSheetBlock)]
    if not data_blocks:
        raise RenderError("xlsx-backtest 至少需要 1 个 data_sheet block")
    for block in data_blocks:
        if block.data_ref not in datasets:
            raise RenderError(f"dataset 未登记：{block.data_ref}")

    wb = Workbook()
    wb.remove(wb.active)
    _notes_sheet(wb, spec.blocks)
    for block in data_blocks:
        _data_sheet(wb, block, datasets[block.data_ref])
    for block in spec.blocks:
        if isinstance(block, MetricsSheetBlock):
            missing = [ref for ref in block.data_refs if ref not in datasets]
            if missing:
                raise RenderError(f"dataset 未登记：{missing}")
            _metrics_sheets(wb, block, datasets)
    for i, block in enumerate(spec.blocks):
        if isinstance(block, TableBlock):
            _table_sheet(wb, block, i)
    _evidence_sheet(wb, evidence)

    props = wb.properties
    props.title = spec.title
    props.description = f"生成于 {datetime.now(UTC).isoformat(timespec='seconds')}（finance-agent）"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
